import asyncio
import discord
import time
import random
import sqlite3
import hashlib
import os
import openpyxl
import discord.enums
import ast
import re
import gensim
import matplotlib as mpl
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from discord.ui import Button, View
from discord import app_commands
from discord.ext import commands
from discord.utils import get
from discord.http import Route
from discord import Intents
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from discord.ext import tasks
from datetime import datetime
from sklearn.feature_extraction.text import CountVectorizer
from nltk import ngrams
from gensim.models import Word2Vec
from collections import Counter
from collections import defaultdict
from sklearn.manifold import TSNE
import matplotlib as mpl
import matplotlib.pyplot as plt
import pandas as pd

class CostomCommandTree(app_commands.CommandTree):
    def add_command(self, command, *args, override=True, **kwargs):
        super().add_command(command, *args, override=override, **kwargs)
intents = discord.Intents.all()
intents.guilds = True
client = discord.Client(intents= intents)
openxl = openpyxl.load_workbook("nemobot-3.xlsx")
bot = commands.Bot(command_prefix='!',intents=intents, tree_cls=CostomCommandTree)

dt = datetime.now()
global preset
preset = 0
global start_time
start_time = None


# SQLite 데이터베이스 연결
conn = sqlite3.connect('conversation_data.db')
c = conn.cursor()

# 테이블 생성
c.execute('''
    CREATE TABLE IF NOT EXISTS conversations
    (id INTEGER PRIMARY KEY AUTOINCREMENT, word_index TEXT, introduction TEXT, i_bigrams TEXT, i_trigrams TEXT, answer TEXT, a_bigrams TEXT, a_trigrams TEXT)
''')

# FTS5 테이블 생성
c.execute("CREATE VIRTUAL TABLE IF NOT EXISTS wordserch USING fts5(word_index, content='conversations', content_rowid= 'id')")
# 트리거 생성
c.execute("""
    CREATE TRIGGER IF NOT EXISTS fts_trigger AFTER INSERT ON conversations BEGIN 
        INSERT INTO wordserch (rowid, word_index) 
        VALUES (new.id, new.word_index);
    END;
""")
c.execute("""
    CREATE TRIGGER IF NOT EXISTS fts_delete_trigger AFTER DELETE ON conversations BEGIN 
        DELETE FROM wordserch WHERE rowid = old.id;
    END;
""")


def sheetname(servername):
    if servername in openxl.sheetnames:
        if openxl.active.title != servername:
            openxl.move_sheet(servername,0)
        openxl.save('nemobot.xlsx')
        return servername
                        
    elif servername not in openxl.sheetnames:
        openxl.create_sheet('{}'.format(servername),0)
        openxl.save('nemobot.xlsx')
        return servername

def checkRow(ctx):
        for row in range(1, openxl[sheetname(ctx)].max_row + 1):
            if openxl[sheetname(ctx)].cell(row, 1).value is None or '':
                break
        return row +1
        
def checkCol(ctx):
        for column in range(1, openxl[sheetname(ctx)].max_column+1):
            if openxl[sheetname(ctx)].cell(1, column).value is None:
                break
        return column +1

def sheetsort(ctx):
    sheet = openxl[sheetname(ctx)]
    
    _row = checkRow(ctx)
    _col = checkCol(ctx)
    # 데이터 불러오기
    data = []
    for row_num in range(1, _row):
        row_data = []
        for col_num in range(1, _col):
            cell_value = openxl[sheetname(ctx)].cell(row=row_num, column=col_num).value
            if cell_value is not None:
                row_data.append(cell_value)
        data.append(row_data)
    
    # B열 기준 내림차순 정렬
    sorted_data = sorted(data, key=lambda x: x[1], reverse=True)
    
    # 다시 셀에 입력
    for row_idx, row_data in enumerate(sorted_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    openxl.save("nemobot.xlsx")

def editbnr(ctx):
    servername = sheetname(ctx.guild.name)
    editbanner = discord.Embed(title="출석체크!",description="순위", colour=discord.Colour.blue())
    editbanner.set_footer(text="출석하려면 버튼을 눌러주세요")
    sheetsort(servername)#시트소팅    
    
    _row = checkRow(servername)
    for row in range(1, _row):
         value2 = "누적 {} 회 오늘은 {}\n마지막 출석 {}".format (openxl[servername].cell(row, 2).value, openxl[servername].cell(row, 3).value, openxl[servername].cell(row, 4).value)
         editbanner.add_field(name=openxl[servername].cell(row, 1).value, value=value2, inline=False)
     
    return editbanner


@bot.event
async def on_ready():
    global bannersend
    print('다음으로 로그인합니다: ')
    print(bot.user.name)
    print(client.shard_id)
    print(client.application_id)
    print('connection was succesful')
    await bot.change_presence(status=discord.Status.online, activity=discord.Game("/명령어"))
    for joiningguild in bot.guilds:
        if f'{bot.user.name.replace(" ", "-").replace("(","-").replace(")","")}의-학습-허용' not in [channel.name for channel in joiningguild.text_channels]:
            role = discord.utils.get(joiningguild.roles, name=bot.user.name)
            overwrites = {
                joiningguild.default_role: discord.PermissionOverwrite(send_messages=False),
                role: discord.PermissionOverwrite(send_messages=True)
            }
            category = discord.utils.get(joiningguild.categories, name = bot.user.name)
            print(f"debugging category: {category}")
            if category:
                pass
            else:
                await joiningguild.create_category(bot.user.name)
            category = discord.utils.get(joiningguild.categories, name = bot.user.name)
            if f'{bot.user.name.replace(" ", "-").replace("(","-").replace(")","")}의-학습-허용' not in [channel.name for channel in joiningguild.text_channels]:
                await joiningguild.create_text_channel(f'{bot.user.name.replace(' ', '-').replace('(','-').replace(')','')}의-학습-허용', category = category, overwrites=overwrites, topic ="채팅학습 허용을 선택합니다")


        print("bot is checking channel:")
        
        for ch in joiningguild.text_channels:
            print(ch.name)
            if ch.name == f"{bot.user.name.replace(" ","-").replace("(","-").replace(")","")}출석체크방-beta":    
                target_message=await ch.history(limit=None).flatten()
                print("bot is checking message:")
                print([target_message])
                await ch. delete_messages(target_message)
                value1 ,value2 = await bannersendfnc(ch)
                await ch.send("봇이 재연결되었습니다", delete_after = 2)
                global bannersend
                bannersend=await ch.send(embed=value1, view= value2)
                break
        else:
            pass
            
    while on_ready:
        res = await bot.wait_for('interaction')
        for item in [res.data]:
            try: 
                if item['custom_id'] == 'checkbutton':
                    # 'checkbutton' 뒤의 문자열 추출
                    channel_name = item['custom_id'][len('checkbutton'):].strip("', ")
                    print("{}채널에서 버튼이 눌렸습니다".format(channel_name))
                    inputres = res
                    print(inputres)              
                    await checkbutton_callback(inputres)
                
                else:
                    print([res.data])
            except:
                pass
                
                                                          
@bot.event
async def bannersendfnc(channel):
    checkbutton = Button(label="출석확인", style=discord.ButtonStyle.green, custom_id= 'checkbutton')
    
    view = View()
    view.add_item(checkbutton)

    es = discord.Embed(title="출석체크!",description="순위", colour=discord.Colour.blue())
    es.set_footer(text="출석하려면 버튼을 눌러주세요")
    ctx = channel.guild.name
    servername = sheetname(ctx)
    _row = checkRow(ctx)  
    for row in range(1, _row):
         if openxl[servername].cell(row, 4).value != "{}년 {}월 {}일".format (datetime.now().year, datetime.now().month, datetime.now().day):
             openxl[servername].cell(row= row, column=3, value = '출석안함')
         
    for row in range(1, _row):
         value2 = "누적 {} 회 오늘은 {}\n마지막 출석 {}".format (openxl[servername].cell(row, 2).value, openxl[servername].cell(row, 3).value, openxl[servername].cell(row, 4).value)
         es.add_field(name=openxl[servername].cell(row, 1).value, value=value2, inline=False) 
    
    return es, view
         
@tasks.loop(seconds=10)
async def dailyset(ctx):
       if (dt.hour == 0) and (dt.minute == 0) and (dt.second > 0):
           servername = sheetname(ctx.guild.name)
           _row=checkRow(servername)
           for row in range(1, _row):
               openxl[servername].cell(row=row, column=3,value='출석안함')
           
           es = discord.Embed(title="출석체크!",description="순위", colour=discord.Colour.blue())
           es.set_footer(text="출석하려면 버튼을 눌러주세요")
        
           _row = checkRow(servername)
           for row in range(1, _row):
                 value2 = "누적 {} 회 오늘은 {}\n마지막 출석 {}".format (openxl[servername].cell(row, 2).value, openxl[servername].cell(row, 3).value, openxl[servername].cell(row, 4).value)
                 es.add_field(name=openxl[servername].cell(row, 1).value, value=value2, inline=False)
           for joiningguild in bot.guilds:
               includech = joiningguild.text_channels
               for ch in includech:
                   if ch.name == f"{bot.user.name.replace(" ","-").replace("(","-").replace(")","")}출석체크방-beta":
                       target_message=await ch.history(limit=1).flatten()
                       await target_message[0].edit(embed=es)                
                       break
               else:
                   return
           else:
               return
           await asyncio.sleep(3540)
                 
@bot.command()
async def role(ctx):
    print(await ctx.guild.fetch_roles())
    a = ctx.guild.get_member(bot.user.id).roles
    await ctx.reply(f"제가 가진 역할은{a}")

@bot.command()
async def test1(ctx):
    await ctx.reply(f"test 성공하셧네요",)


@bot.tree.command(name='test', description='testing tress command')
async def test(ctx):
    await ctx.reply(f"test 성공하셧네요")


@app_commands.command()
async def fruits(interaction: discord.Interaction, fruit: str):
    await interaction.response.send_message(f'Your favourite fruit seems to be {fruit}')

@fruits.autocomplete('fruit')
async def fruits_autocomplete(
    interaction: discord.Interaction,
    current: str,
) -> list[app_commands.Choice[str]]:
    fruits = ['Banana', 'Pineapple', 'Apple', 'Watermelon', 'Melon', 'Cherry']
    return [
        app_commands.Choice(name=fruit, value=fruit)
        for fruit in fruits if current.lower() in fruit.lower()
    ]


@bot.command()
@commands.has_permissions(administrator = True)
async def removecommand(ctx, command):
    for cmd in bot.walk_commands():
        if cmd.name.lower() == command.lower():
            bot.remove_command(cmd)
            await ctx.message.add_reaction('✅')
            return
    await ctx.channel.send(f"Command `{command}` does not exist")

@bot.tree.command(name="setup",description="setup attendent")
async def setup(ctx):
    servername = sheetname(ctx.guild.name)
    if ctx.user.guild_permissions.administrator:
        _row=checkRow(servername)
        for row in range(1, _row):
            openxl[servername].cell(row=row, column=3,value='출석안함')
            
        sheetsort(servername)#시트소트
        await ctx.reply("dailyset 출석안함 done")
        
        es = discord.Embed(title="출석체크!",description="순위", colour=discord.Colour.blue())
        es.set_footer(text="출석하려면 버튼을 눌러주세요")
        
        _row = checkRow(servername)
        for row in range(1, _row):
                 value2 = "누적 {} 회 오늘은 {}\n마지막 출석 {}".format (openxl[servername].cell(row, 2).value, openxl[servername].cell(row, 3).value, openxl[servername].cell(row, 4).value)
                 es.add_field(name=openxl[servername].cell(row, 1).value, value=value2, inline=False)
        
        includech = ctx.guild.text_channels
        for ch in includech:
            if ch.name == f"{bot.user.name.replace(" ","-").replace("(","-").replace(")","")}출석체크방-beta":
                        target_message=await ch.history(limit=1).flatten()
                        await target_message[0].edit(embed=es) 
                        break
        else:
            ctx.send("전용채널이 없습니다")
    
    else:
        ctx.reply("관리자 권한이 필요합니다",ephemeral=True)
        
@bot.tree.command(name="전용채널생성",description="전용채널을 생성합니다")
async def 전용채널생성(ctx): 
    if ctx.channel.name != f"{bot.user.name.replace(" ","-").replace("(","-").replace(")","")}출석체크방-beta" and ctx.user.guild_permissions.administrator:
        
        role = discord.utils.get(ctx.guild.roles, name=bot.user.name)
        overwrites = {
                ctx.guild.default_role: discord.PermissionOverwrite(send_messages=False),
                role: discord.PermissionOverwrite(send_messages=True)
            }
        
        await ctx.reply("명령어 사용!")
        
        if not f'{bot.user.name.replace(' ', '-').replace('(','-').replace(')','')}의-학습-허용' in ctx.guild.text_channels:
            if not f'{bot.user.name}' in ctx.guild.categories:
                await ctx.guild.create_category(bot.user.name)
            else:
                pass
            category = discord.utils.get(ctx.guild.categories, name = bot.user.name)
            await ctx.guild.create_text_channel(f'{bot.user.name.replace(' ', '-').replace('(','-').replace(')','')}의-학습-허용', category = category, overwrites=overwrites, topic ="채팅학습 허용을 선택합니다")

        h = False
        t = False
        
        includech = ctx.guild.text_channels
        for ch in includech:
                if ch.name == f'{bot.user.name}출석체크방-beta':
                    await ctx.send(embed=discord.Embed(title="전용채널이 존재합니다",description=ch.name, color=0xff0000))
                    h = True
                    break
                
            
        includect = ctx.guild.categories
        for ct in includect:
                if ct.name == bot.user.name:
                    await ctx.send(embed=discord.Embed(title="카테고리가 존재합니다",description=ct.name, color= 0xff0000))
                    t = True
                    break
            
        if h == False and t == True:
            category = discord.utils.get(ctx.guild.categories, name = bot.user.name)
            await ctx.guild.create_text_channel(f'{bot.user.name}출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")
            await ctx.send(embed=discord.Embed(title="채널을 생성했습니다",description=f"{bot.user.name.replace(" ","-").replace("(","-").replace(")","")}출석체크방-beta", color= 0xff000))
        
        elif h == False and t == False:
            await ctx.guild.create_category(bot.user.name)
            
            category = discord.utils.get(ctx.guild.categories, name = bot.user.name)
            await ctx.guild.create_text_channel(f'{bot.user.name}출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")
            
            embed = discord.Embed(title = "채널 및 카테고리를 생성했습니다", color = 0xff000)        
            embed.add_field(name = "생성된 카테고리", value = "bot.user.name")
            embed.add_field(name = "생성된 채널", value = f"{bot.user.name.replace(" ","-").replace("(","-").replace(")","")}출석체크방-beta")
            await ctx.send(embed = embed)
        
        elif h ==  True and t == False:
            ch = discord.utils.get(ctx.guild.channels, name = f'{bot.user.name}출석체크방-beta')
            await ch.delete()
            category = await ctx.guild.create_category(bot.user.name)
            await ctx.guild.create_text_channel(f'{bot.user.name}출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다n(※알림 꺼두세요※)")          
            await ctx.send(embed=discord.Embed(title="투표 채널을 초기화하고 카테고리를 생성했습니다",description="bot.user.name", color= 0xffff00))
        
        elif h == True and t == True:
            ch = discord.utils.get(ctx.guild.channels, name = f'{bot.user.name}출석체크방-beta')
            await ch.delete()
            category = discord.utils.get(ctx.guild.categories, name = bot.user.name)
            await ctx.guild.create_text_channel(f'{bot.user.name}출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")
            await ctx.send(embed=discord.Embed(title="투표채널을 초기화하였습니다",description=f"{bot.user.name.replace(" ","-").replace("(","-").replace(")","")}출석체크방-beta", color= 0xffff00))
            
        
        await ctx.send("투표가 생성되니다") 
        await ctx.send("투표가 생성된 방의 알림을 꺼주세요")
        
    
    elif ctx.user.guild_permissions.administrator is False:
            await ctx.reply("관리자가 아닙니다",ephemeral=True)
            
    else:
        await ctx.reply("현재 카테고리 제외 다른 채널에 입력해주세요",ephemeral=True)
     

@bot.tree.command(name="전용채널삭제",description="전용채널을 삭제합니다")
async def 전용채널삭제(ctx):
 if ctx.channel.name != f"{bot.user.name.replace(" ","-").replace("(","-").replace(")","")}출석체크방-beta" and ctx.user.guild_permissions.administrator:
        await ctx.reply("명령어 사용!")
        d = 0
        
        includedch =  ctx.guild.text_channels
        for ch in includedch:
            if ch.name == f'{bot.user.name}출석체크방-beta':
                     deletech = discord.utils.get(ctx.guild.channels, name = f'{bot.user.name}출석체크방-beta')
                     await deletech.delete()
                     await ctx.send(embed=discord.Embed (title="채널삭제됨",description=ch.name, color= 0xffff00))
                     d=1
                     
                     
        includedct = ctx.guild.categories
        for ct in includedct:
             if ct.name == bot.user.name:
                     deletect = discord.utils.get(ctx.guild.categories, name = bot.user.name)
                     await deletect.delete()
                     await ctx.send(embed=discord.Embed (title="카테고리 삭제됨",description=ct.name, color= 0xffff00))
                     d=1
        
        if d==0:
                 await ctx.send(embed=discord.Embed (title="삭제할 채널 혹은 카테고리가 존재하지 않습니다",color= 0xff0000))
 
 elif  ctx.channel.name != f"{bot.user.name.replace(" ","-").replace("(","-").replace(")","")}출석체크방-beta" and ctx.user.guild_permissions.administrator is False:
     ctx.reply("관리자 권한이 필요합니다",ephemeral=True)
 
 else:
    await ctx.send("현재 카테고리를 제외한 다른 채널에 입력해주세요",ephemeral=True)
                                                                
                                   
@bot.event
async def on_guild_join(guild): 
    h = False
    t = False

    role = discord.utils.get(guild.roles, name=bot.user.name)
    overwrites = {
            guild.default_role: discord.PermissionOverwrite(send_messages=False),
            role: discord.PermissionOverwrite(send_messages=True)
        }

    
    includech = guild.text_channels
    for ch in includech:
              if ch.name == f'{bot.user.name}출석체크방-beta':                 
                  h = True
                  break
              
         
    includect = guild.categories
    for ct in includect:
              if ct.name == bot.user.name:
                  t = True
                  break
        
    if h == False and t == True:
          category = discord.utils.get(guild.categories, name = bot.user.name)
          await guild.create_text_channel(f'{bot.user.name}출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")
          
      
    elif h == False and t == False:
          await guild.create_category(bot.user.name)
          
          category = discord.utils.get(guild.categories, name = bot.user.name)
          await guild.create_text_channel(f'{bot.user.name}출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")
          
          
      
    elif h ==  True and t == False:
          ch = discord.utils.get(guild.channels, name = f'{bot.user.name}출석체크방-beta')
          await ch.delete()
          category = await guild.create_category(bot.user.name)
          await guild.create_text_channel(f'{bot.user.name}출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")          
          
      
    elif h == True and t == True:
           ch = discord.utils.get(guild.channels, name = f'{bot.user.name}출석체크방-beta')
           await ch.delete()
           category = discord.utils.get(guild.categories, name = bot.user.name)
           await guild.create_text_channel(f'{bot.user.name}출석체크방-beta', category = category, overwrites=overwrites, topic ="출석체크방입니다\n(※알림 꺼두세요※)")
    
    category = discord.utils.get(guild.categories, name = bot.user.name)
    if not f'{bot.user.name.replace(' ', '-').replace('(','-').replace(')','')}의-학습-허용' in includech:
        await guild.create_text_channel(f'{bot.user.name.replace(' ', '-').replace('(','-').replace(')','')}의-학습-허용', category = category, overwrites=overwrites, topic ="채팅학습 허용을 선택합니다")



@bot.event
async def on_guild_channel_create(channel):
    ctx=channel.guild
    if channel.name==f"{bot.user.name.replace(" ","-").replace("(","-").replace(")","")}출석체크방-beta":
        value1,value2 = await bannersendfnc(channel)
        global bannersend
        bannersend= await channel.send(embed=value1, view= value2)
        dailyset.start(ctx)
    if channel.name == f'{bot.user.name.replace(' ', '-').replace('(','-').replace(')','')}의-학습-허용':
        image = discord.File("attendencecheckbot(ai).png", filename="ailogo.png")
        allowembed = discord.Embed(title="현재 해당 봇이 자신의 채팅을 학습하는데 동의하면 아래 ✅에 체크해주세요", description="동의에 체크된 사람의 채팅만 학습하며 체크를 해제할 시 해당 사용자의 채팅 학습을 중단합니다", color=0x00ff56)
        allowembed.set_thumbnail(url="attachment://ailogo.png")
        allowembed.add_field(name="-유의사항-", value= "1. 체크를 해제할 시 이전에 학습한 내용은 삭제되지 않습니다\n2. ✅에 확인되어있는 동안의 채팅만 학습됩니다\n3. 텍스트채팅만 학습하며 해당 봇의 학습이외의 목적으로 사용되지 않습니다 \n4.해당 모델은 디스코드에서 허용된 채팅으로만 데이터를 수집합니다", inline=True)
        permitioncheck = await channel.send(embed=allowembed, file=image)
        await permitioncheck.add_reaction('✅')



togletag = None
convtrance = False
needsetup = False

def is_introduction(tag):
    global togletag, convtrance, conversations , needsetup
    needsetup = False
    # print(f"togletag: {togletag}")
    if convtrance == False:
        if togletag == tag or togletag == None:
            togletag = tag
            # print("1")
            return 'introduction'
        else: #togletag != tag
            togletag = tag
            convtrance = True
            # print("2")
            return 'answer'
        
    elif convtrance == True:
        if togletag == tag:
            # print("3")
            return 'answer'
        else: #togletag != tag
            convtrance = False
            togletag = tag
            needsetup = True
            print("4, needsetup is true now")
            return 'introduction'


def calculate_freq(raw_message):
    words = raw_message.split()

    # 단어, 바이그램, 트라이그램 빈도 업데이트
    word_freq = dict(Counter(words))

    # bi-gram 생성 및 빈도 계산
    bigrams = list(ngrams(words, 2)) if len(words) >= 2 else []
    bigram_freq = dict(Counter(bigrams))

    # tri-gram 생성 및 빈도 계산
    trigrams = list(ngrams(words, 3)) if len(words) >= 3 else []
    trigram_freq = dict(Counter(trigrams))

    # print(word_freq, bigram_freq, trigram_freq)

    return word_freq, bigram_freq, trigram_freq




def handle_none(value):
    return '' if value is None else value


conversations = defaultdict(lambda: {"introduction": [], "answer": []}) 

def save_data(tag, message):
    global conversations, word_freq,  bigram_freq, trigram_freq, needsetup, channel_id
    conversationident = is_introduction(tag)

    if needsetup == True:
        raw_intoduction = ' '.join([m['raw_message'] for m in conversations['currunt_conversation']["introduction"]])
        word_freq, bigram_freq, trigram_freq = calculate_freq(raw_intoduction)
        conversations['currunt_conversation']['introduction'].append({"words": word_freq, "Bi-grams": bigram_freq, "Tri-grams": trigram_freq})
        raw_answer = ' '.join([m['raw_message'] for m in conversations['currunt_conversation']["answer"]])
        word_freq, bigram_freq, trigram_freq = calculate_freq(raw_answer)
        conversations['currunt_conversation']['answer'].append({"words": word_freq, "Bi-grams": bigram_freq, "Tri-grams": trigram_freq})
        # print(f"final conversation is : {conversations}")

        introduction = conversations['currunt_conversation']["introduction"][1]['words']
        i_bigrams = conversations['currunt_conversation']["introduction"][1]['Bi-grams']
        i_trigrams = conversations['currunt_conversation']["introduction"][1]['Tri-grams']
        answer = conversations['currunt_conversation']["answer"][1]['words']
        a_bigrams = conversations['currunt_conversation']["answer"][1]['Bi-grams']
        a_trigrams = conversations['currunt_conversation']["answer"][1]['Tri-grams']

        # print(f"this is the final conversation data: {conversations}")
        
        # introduction과 answer의 단어들을 문자열로 변환
        introduction_str = ' '.join(conversations['currunt_conversation']["introduction"][1]['words'])
        answer_str = ' '.join(conversations['currunt_conversation']["answer"][1]['words'])
        word_index = introduction_str + ' ' + answer_str

        # 단어 검색 및 id 반환
        targetserch = re.sub(r'\W+', ' ', word_index)
        targets = targetserch.split()
        # print(f"targets: {targets}")
        datanotfound = None
        all_ids = []
        candidates = {}
        for target in targets:
            # print(f"target: {target}")
            c.execute(f"SELECT rowid FROM wordserch WHERE word_index MATCH '{target}'")
            ids = c.fetchall()
            ids = [id[0] for id in ids]
            all_ids.extend(ids)
        all_ids = list(set(all_ids))

        if len(all_ids)==0:
            if datanotfound == None:
                datanotfound = True
            else:
                datanotfound = True
        else:
            datanotfound = False
            
            # 해당하는 항목이 있다면 기존 항목 수정
            for id in all_ids:
                c.execute(f"SELECT * FROM conversations WHERE id = '{id}'")
                id_data = c.fetchone()
                # print(f"id row data: {id_data}")
                candidate = {}
                candidate['id'], candidate['word_index'], candidate['introduction'], candidate['i_bigrams'], candidate['i_trigrams'], candidate['answer'], candidate['a_bigrams'], candidate['a_trigrams'] = id_data
               

                # candidate 딕셔너리의 데이터를 str -> ast로 변환
                candidate_i_bigrams = ast.literal_eval(candidate['i_bigrams'])
                candidate_i_trigrams = ast.literal_eval(candidate['i_trigrams'])
                candidate_a_bigrams = ast.literal_eval(candidate['a_bigrams'])
                candidate_a_trigrams = ast.literal_eval(candidate['a_trigrams'])
                candidate_introduction = ast.literal_eval(candidate['introduction'])
                candidate_answer = ast.literal_eval(candidate['answer'])
                new_i_bigrams = ast.literal_eval(str(i_bigrams))
                new_i_trigrams = ast.literal_eval(str(i_bigrams))
                new_a_bigrams = ast.literal_eval(str(a_bigrams))
                new_a_trigrams = ast.literal_eval(str(a_trigrams))
                new_introduction = ast.literal_eval(str(introduction))
                new_answer = ast.literal_eval(str(answer))

                trigrampass = True
                bigrampass = True   
                # trigram 확인
                if set(new_i_trigrams.keys()).intersection(set(candidate_i_trigrams.keys())):
                    # 워드
                    for word, count in new_introduction.items():
                        if word in candidate_introduction:
                            candidate_introduction[word] += count
                        else:
                            candidate_introduction[word] = count
                    # 바이그램
                    for bigram, count in new_i_bigrams.items():
                        if bigram in candidate_i_bigrams:
                            candidate_i_bigrams[bigram] += count
                        else:
                            candidate_i_bigrams[bigram] = count
                    #트라이그램
                    for trigram, count in new_i_trigrams.items():
                        if trigram in candidate_i_trigrams:
                            candidate_i_trigrams[trigram] += count
                        else:
                            candidate_i_trigrams[trigram] = count
                    trigrampass = False
                    # print("found introduction matching whith introduction data at trigram")
                if set(new_a_trigrams.keys()).intersection(set(candidate_i_trigrams.keys())):
                    #워드
                    for word, count in new_answer.items():
                        if word in candidate_introduction:
                            candidate_introduction[word] += count
                        else:
                            candidate_introduction[word] = count
                    # 바이그램
                    for bigram, count in new_a_bigrams.items():
                        if bigram in candidate_i_bigrams:
                            candidate_i_bigrams[bigram] += count
                        else:
                            candidate_i_bigrams[bigram] = count
                    # 트라이그램
                    for trigram, count in new_a_trigrams.items():
                        if trigram in candidate_i_trigrams:
                            candidate_i_trigrams[trigram] += count
                        else:
                            candidate_i_trigrams[trigram] = count
                    trigrampass = False
                    # print("found answer matching with introduction data as trigram")      
                if set(new_i_trigrams.keys()).intersection(set(candidate_a_trigrams.keys())):
                    # 워드
                    for word, count in new_introduction.items():
                        if word in candidate_answer:
                            candidate_answer[word] += count
                        else:
                            candidate_answer[word] = count
                    # 바이그램
                    for bigram, count in new_i_bigrams.items():
                        if bigram in candidate_a_bigrams:
                            candidate_a_bigrams[bigram] += count
                        else:
                            candidate_a_bigrams[bigram] = count
                    #트라이그램
                    for trigram, count in new_i_trigrams.items():
                        if trigram in candidate_a_trigrams:
                            candidate_a_trigrams[trigram] += count
                        else:
                            candidate_a_trigrams[trigram] = count
                    trigrampass = False
                    # print("found introduction matching whith answer data at trigram")
                if set(new_a_trigrams.keys()).intersection(set(candidate_a_trigrams.keys())):
                    # 워드
                    for word, count in new_answer.items():
                        if word in candidate_answer:
                            candidate_answer[word] += count
                        else:
                            candidate_answer[word] = count
                    # 바이그램
                    for bigram, count in new_a_bigrams.items():
                        if bigram in candidate_a_bigrams:
                            candidate_a_bigrams[bigram] += count
                        else:
                            candidate_a_bigrams[bigram] = count
                    #트라이그램
                    for trigram, count in new_a_trigrams.items():
                        if trigram in candidate_a_trigrams:
                            candidate_a_trigrams[trigram] += count
                        else:
                            candidate_a_trigrams[trigram] = count
                    trigrampass = False
                    # print("found answer matching whith answer data at trigram")

                # bigram 확인
                if trigrampass == True:
                    if set(new_i_bigrams.keys()).intersection(set(candidate_i_bigrams.keys())):
                        # 워드
                        for word, count in new_introduction.items():
                            if word in candidate_introduction:
                                candidate_introduction[word] += count
                            else:
                                candidate_introduction[word] = count
                        # 바이그램
                        for bigram, count in new_i_bigrams.items():
                            if bigram in candidate_i_bigrams:
                                candidate_i_bigrams[bigram] += count
                            else:
                                candidate_i_bigrams[bigram] = count
                        #트라이그램
                        for trigram, count in new_i_trigrams.items():
                            if trigram in candidate_i_trigrams:
                                candidate_i_trigrams[trigram] += count
                            else:
                                candidate_i_trigrams[trigram] = count
                        bigrampass = False
                        # print("found introduction matching whith introduction data at bigram")
                    if set(new_a_bigrams.keys()).intersection(set(candidate_i_bigrams.keys())):
                        #워드
                        for word, count in new_answer.items():
                            if word in candidate_introduction:
                                candidate_introduction[word] += count
                            else:
                                candidate_introduction[word] = count
                        # 바이그램
                        for bigram, count in new_a_bigrams.items():
                            if bigram in candidate_i_bigrams:
                                candidate_i_bigrams[bigram] += count
                            else:
                                candidate_i_bigrams[bigram] = count
                        # 트라이그램
                        for trigram, count in new_a_trigrams.items():
                            if trigram in candidate_i_trigrams:
                                candidate_i_trigrams[trigram] += count
                            else:
                                candidate_i_trigrams[trigram] = count
                        bigrampass = False
                        # print("found answer matching with introduction data as bigram")
                    if set(new_i_bigrams.keys()).intersection(set(candidate_a_bigrams.keys())):
                        # 워드
                        for word, count in new_introduction.items():
                            if word in candidate_answer:
                                candidate_answer[word] += count
                            else:
                                candidate_answer[word] = count
                        # 바이그램
                        for bigram, count in new_i_bigrams.items():
                            if bigram in candidate_a_bigrams:
                                candidate_a_bigrams[bigram] += count
                            else:
                                candidate_a_bigrams[bigram] = count
                        #트라이그램
                        for trigram, count in new_i_trigrams.items():
                            if trigram in candidate_a_trigrams:
                                candidate_a_trigrams[trigram] += count
                            else:
                                candidate_a_trigrams[trigram] = count
                        bigrampass = False
                        # print("found introduction matching whith answer data at bigram")
                    if set(new_a_bigrams.keys()).intersection(set(candidate_a_bigrams.keys())):
                        # 워드
                        for word, count in new_answer.items():
                            if word in candidate_answer:
                                candidate_answer[word] += count
                            else:
                                candidate_answer[word] = count
                        # 바이그램
                        for bigram, count in new_a_bigrams.items():
                            if bigram in candidate_a_bigrams:
                                candidate_a_bigrams[bigram] += count
                            else:
                                candidate_a_bigrams[bigram] = count
                        #트라이그램
                        for trigram, count in new_a_trigrams.items():
                            if trigram in candidate_a_trigrams:
                                candidate_a_trigrams[trigram] += count
                            else:
                                candidate_a_trigrams[trigram] = count
                        bigrampass = False
                        # print("found answer matching with answer data at bigram")

                # 단어 확인
                if bigrampass == True:                           
                    if set(introduction_str.split()).intersection(candidate_introduction):
                        # 워드
                        for word, count in new_introduction.items():
                            if word in candidate_introduction:
                                candidate_introduction[word] += count
                            else:
                                candidate_introduction[word] = count
                        # 바이그램
                        for bigram, count in new_i_bigrams.items():
                            if bigram in candidate_i_bigrams:
                                candidate_i_bigrams[bigram] += count
                            else:
                                candidate_i_bigrams[bigram] = count
                        #트라이그램
                        for trigram, count in new_i_trigrams.items():
                            if trigram in candidate_i_trigrams:
                                candidate_i_trigrams[trigram] += count
                            else:
                                candidate_i_trigrams[trigram] = count
                        # print("found introduction matching whith introduction data at words")
                    if set(answer_str.split()).intersection(candidate_introduction):
                        #워드
                        for word, count in new_answer.items():
                            if word in candidate_introduction:
                                candidate_introduction[word] += count
                            else:
                                candidate_introduction[word] = count
                        # 바이그램
                        for bigram, count in new_a_bigrams.items():
                            if bigram in candidate_i_bigrams:
                                candidate_i_bigrams[bigram] += count
                            else:
                                candidate_i_bigrams[bigram] = count
                        # 트라이그램
                        for trigram, count in new_a_trigrams.items():
                            if trigram in candidate_i_trigrams:
                                candidate_i_trigrams[trigram] += count
                            else:
                                candidate_i_trigrams[trigram] = count
                        # print("found answer matching with introduction data as words")
                    if set(introduction_str.split()).intersection(candidate_answer):
                        # 워드
                        for word, count in new_introduction.items():
                            if word in candidate_answer:
                                candidate_answer[word] += count
                            else:
                                candidate_answer[word] = count
                        # 바이그램
                        for bigram, count in new_i_bigrams.items():
                            if bigram in candidate_a_bigrams:
                                candidate_a_bigrams[bigram] += count
                            else:
                                candidate_a_bigrams[bigram] = count
                        #트라이그램
                        for trigram, count in new_i_trigrams.items():
                            if trigram in candidate_a_trigrams:
                                candidate_a_trigrams[trigram] += count
                            else:
                                candidate_a_trigrams[trigram] = count
                        # print("found introduction matching whith answer data at words")
                    if set(answer_str.split()).intersection(candidate_answer):
                        for word, count in new_answer.items():
                            if word in candidate_answer:
                                candidate_answer[word] += count
                            else:
                                candidate_answer[word] = count
                        # 바이그램
                        for bigram, count in new_a_bigrams.items():
                            if bigram in candidate_a_bigrams:
                                candidate_a_bigrams[bigram] += count
                            else:
                                candidate_a_bigrams[bigram] = count
                        #트라이그램
                        for trigram, count in new_a_trigrams.items():
                            if trigram in candidate_a_trigrams:
                                candidate_a_trigrams[trigram] += count
                            else:
                                candidate_a_trigrams[trigram] = count
                        # print("found answer matching with answer data at words")
                                
                # 딕셔너리 리스트
                dicts = [candidate_i_bigrams, candidate_i_trigrams, candidate_a_bigrams, candidate_a_trigrams]

                # 키 튜플의 모든 요소를 리스트로 만들기
                words_in_common = [item for dict_example in dicts for key in dict_example.keys() for item in key]

                # targets와 word_index 사이 차집합 계산
                word_index_set = set(candidate['word_index'].split())
                targets_set = set(targets)
                commonword = set(words_in_common)
                new_words = targets_set.intersection(commonword) - word_index_set
                # 새로운 단어를 word_index에 추가
                candidate['word_index'] += ' ' + ' '.join(new_words)
                candidate['i_bigrams'] = str(candidate_i_bigrams)
                candidate['i_trigrams'] = str(candidate_i_trigrams)
                candidate['a_bigrams'] = str(candidate_a_bigrams)
                candidate['a_trigrams'] = str(candidate_a_trigrams)
                candidate['introduction'] = str(candidate_introduction)
                candidate['answer'] = str(candidate_answer)
                # print(f"candidate: {candidate}")

                
                # 딕셔너리 저장
                print(f"candidateid:{candidate["id"]}")
                candidates[candidate['id']] = candidate
            # print(f"candidates: {candidates}")
        for kid in candidates.keys():
            # print(kid)
            c.execute(f"""
                        UPDATE conversations 
                        SET word_index = ?, introduction = ?, i_bigrams = ?, i_trigrams = ?, answer = ?, a_bigrams = ?, a_trigrams = ?
                        WHERE id = '{kid}'
                    """, (candidate['word_index'], candidate['introduction'], candidate['i_bigrams'], candidate['i_trigrams'], candidate['answer'], candidate['a_bigrams'], candidate['a_trigrams']))
            conn.commit()
        if datanotfound == True:
            # 해당하는 항목이 없을시 새로울 항목으로 추가
                c.execute("INSERT INTO conversations (word_index, introduction, i_bigrams, i_trigrams, answer, a_bigrams, a_trigrams) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (word_index, str(introduction), str(i_bigrams), str(i_trigrams), str(answer), str(a_bigrams), str(a_trigrams)))
                conn.commit
        else:
            pass
            
        # 변경 사항 커밋
        # 데이터 삽입
        conn.commit()
        conversations = defaultdict(lambda: {"introduction": [], "answer": []})

    # raw_message를 계속 이어붙임
    if len(conversations['currunt_conversation'][f"{conversationident}"]) > 0:
        conversations['currunt_conversation'][f"{conversationident}"][-1]['raw_message'] += ' ' + message
        # print(conversations)
    else:
        conversations['currunt_conversation'][f"{conversationident}"].append({"raw_message": message})
    print(f"cnversations: {conversations}")



@bot.event
async def on_message(ctx):
    # # print([ctx])
    # print(ctx.content)
    # print([ctx.type])
    if(ctx.type[1] == 0 and len(ctx.content)>0 and ctx.author.bot != True and (ctx.content[0].isalnum() or ctx.content[0] == '.') and not ctx.mentions and not 'http://' in ctx.content and not 'https://' in ctx.content):
        if random.randrange(1,101) == 7:
            # urllib.request.urlretrieve("https://www", ".png")
            image = discord.File("gotcha1.png", filename="gotcha.png")
            embed = discord.Embed(title="**당신의 메세지는 1%의 확률을 뚫고 봇의 가챠본능 이스터에그를 확인했습니다**", description="보상으로 가챠코드를 보여드리죠", color=0x00ff56)
            embed.set_image(url="attachment://gotcha.png")
            embed.add_field(name="굉장하군요", value= "축하드립니다", inline=True)
            await ctx.reply(embed=embed, file=image)
        sentence = ctx.content
        tag = hashlib.md5(str(ctx.author.id).encode()).hexdigest()  # ID 생성
        targetch = discord.utils.get(ctx.guild.channels, name = f"{bot.user.name.replace(" ","-").replace("(","-").replace(")","")}의-학습-허용")
        if targetch:
            async for message in targetch.history():
                if message.author.id == bot.user.id:
                    for reaction in message.reactions:
                        if str(reaction.emoji) == '✅':
                            # 학습 확인 사용자 가져옴
                            users = [user.id async for user in reaction.users()]
                            if ctx.author.id in users:
                                print("saving data")
                                save_data(tag, sentence)
                            else:
                                print("cannotuse data")
                                save_data(tag, "")
    elif ctx.author.id == bot.user.id:
        tag = hashlib.md5(str(ctx.author.id).encode()).hexdigest()  # ID 생성
        save_data(tag, ctx.content)

    model_path = "chat model"
    if os.path.isfile(model_path):
        model = gensim.models.Word2Vec.load(model_path)

    mention = [member.id for member in ctx.mentions]
    if (random.randrange(1,2) == 1 or (bot.user.id in mention)) and os.path.isfile(model_path) and ctx.type[1] == 0 and ctx.author.bot != True and (ctx.content[0].isalnum() or ctx.content[0] == '.' or ctx.content[0] == '@') and not 'http://' in ctx.content and not 'https://' in ctx.content:
        print("generating answer")
        word_dict, bigram_dict, trigram_dict = calculate_freq(re.sub(r'\W+', ' ', ctx.content))
        word_list = list(word_dict.keys())
        bigram_list = list(bigram_dict.keys())
        trigram_list = list(trigram_dict.keys())
        id_list = []
        intersection_i_trigram = []
        intersection_a_trigram = []
        intersection_i_bigram = []
        intersection_a_bigram = []
        id_list_select = {}

        # 해당되는 row id 반환  
        for word in word_list:
            c.execute(f"SELECT rowid FROM wordserch WHERE word_index MATCH '{word}'")
            id_list.extend(item[0] for item in c.fetchall())
        id_list = list(set(id_list))
        print(id_list)

        # 트라이그램 바이그램 단어 순으로 데이터베이스에 데이터가 있는지 확인
        for id in id_list:
            c.execute(f"SELECT i_trigrams, a_trigrams FROM conversations WHERE id = '{id}'")
            result_trigram = c.fetchone()
            c.execute(f"SELECT i_bigrams, a_bigrams FROM conversations WHERE id = '{id}'")
            result_bigram = c.fetchone()
            c.execute(f"SELECT introduction, answer FROM conversations WHERE id = '{id}'")
            result = c.fetchone()
            print(f"result:{result}")
            trigram_exist = False
            bigram_exist = False
            #트라이그램 확인
            if result_trigram:
                i_trigrams, a_trigrams = result_trigram
                print(f"i_trigrams: {i_trigrams}, a_trigrams: {a_trigrams}")
                i_trigrams = ast.literal_eval(i_trigrams)
                a_trigrams = ast.literal_eval(a_trigrams)
                intersection_i_trigram = set(trigram_list).intersection(set(i_trigrams.keys()))
                intersection_a_trigram = set(trigram_list).intersection(set(a_trigrams.keys()))
                print(intersection_i_trigram, intersection_a_trigram)
                # 트라이그램 존재
                if intersection_i_trigram or intersection_a_trigram:
                    trigram_exist = True
                    len_intersection_i_trigram = len(intersection_i_trigram)
                    len_intersection_a_trigram = len(intersection_a_trigram)
                    if len_intersection_i_trigram > len_intersection_a_trigram:
                        print(f"matching {id}, introduction")
                        i_values = [value for key, value in i_trigrams.items() if key in intersection_i_trigram]
                        max_i_value = max(set(i_values))
                        id_list_select[id] = (max_i_value, "trigram", "introduction")

                    elif len_intersection_i_trigram < len_intersection_a_trigram:
                        print(f"matching {id}, answer")
                        a_values = [value for key, value in a_trigrams.items() if key in intersection_a_trigram]
                        max_a_value = max(set(a_values))
                        id_list_select[id] = (max_a_value, "trigram", "answer")

                    else:
                        i_values = [value for key, value in i_trigrams.items() if key in intersection_i_trigram]
                        a_values = [value for key, value in a_trigrams.items() if key in intersection_a_trigram]

                        max_i_value = max(set(i_values))
                        max_a_value = max(set(a_values))

                        if max_a_value < max_i_value:
                            print(f"matching {id}, introduction")
                            id_list_select[id] = (max_i_value, "trigram", "introduction")
                        elif max_a_value > max_i_value:
                            print(f"matching {id}, answer")
                            id_list_select[id] = (max_a_value, "trigram", "answer")
                        else:
                            print(f"matching {id}, introduction, answer")
                            id_list_select[id] = (max_a_value, "trigram", "both")
            
            # 바이그램 확인
            if trigram_exist == False and result_bigram:
                i_bigrams, a_bigrams = result_bigram
                print(f"i_bigrams: {i_bigrams}, a_bigrams: {a_bigrams}")
                i_bigrams = ast.literal_eval(i_bigrams)
                a_bigrams = ast.literal_eval(a_bigrams)
                intersection_i_bigram = set(bigram_list).intersection(set(i_bigrams.keys()))
                intersection_a_bigram = set(bigram_list).intersection(set(a_bigrams.keys()))
                print(intersection_i_bigram, intersection_a_bigram)
                # 바이그램 존재
                if intersection_i_bigram or intersection_a_bigram:
                    bigram_exist = True
                    len_intersection_i_bigram = len(intersection_i_bigram)
                    len_intersection_a_bigram = len(intersection_a_bigram)
                    if len_intersection_i_bigram > len_intersection_a_bigram:
                        print(f"matching {id}, introduction")
                        i_values = [value for key, value in i_bigrams.items() if key in intersection_i_bigram]
                        max_i_value = max(set(i_values))
                        id_list_select[id] = (max_i_value, "bigram", "introduction")

                    elif len_intersection_i_bigram < len_intersection_a_bigram:
                        print(f"matching {id}, answer")
                        a_values = [value for key, value in a_bigrams.items() if key in intersection_a_bigram]
                        max_a_value = max(set(a_values))
                        id_list_select[id] = (max_a_value, "bigram", "answer")

                    else:
                        i_values = [value for key, value in i_bigrams.items() if key in intersection_i_bigram]
                        a_values = [value for key, value in a_bigrams.items() if key in intersection_a_bigram]

                        max_i_value = max(set(i_values))
                        max_a_value = max(set(a_values))

                        if max_a_value < max_i_value:
                            print(f"matching {id}, introduction")
                            id_list_select[id] = (max_i_value, "bigram", "introduction")
                        elif max_a_value > max_i_value:
                            print(f"matching {id}, answer")
                            id_list_select[id] = (max_a_value, "bigram", "answer")
                        else:
                            print(f"matching {id}, introduction, answer")
                            id_list_select[id] = (max_a_value, "bigram", "both")
            
            # 단어 확인
            if bigram_exist == False and result:
                introduction, answer = result
                print(f"introduction: {introduction}, answer: {answer}")
                introduction = ast.literal_eval(introduction)
                answer = ast.literal_eval(answer)
                intersection_introduction = set(word_list).intersection(set(introduction.keys()))
                intersection_answer = set(word_list).intersection(set(answer.keys()))
                print(intersection_introduction, intersection_answer)
                # 단어 존재
                if intersection_introduction or intersection_answer:
                    len_intersection_introduction = len(intersection_introduction)
                    len_intersection_answer = len(intersection_answer)
                    if len_intersection_introduction > len_intersection_answer:
                        print(f"matching {id}, introduction")
                        i_values = [value for key, value in introduction.items() if key in intersection_introduction]
                        max_i_value = max(set(i_values))
                        id_list_select[id] = (max_i_value, "word", "introduction")

                    elif len_intersection_introduction < len_intersection_answer:
                        print(f"matching {id}, answer")
                        a_values = [value for key, value in answer.items() if key in intersection_answer]
                        max_a_value = max(set(a_values))
                        id_list_select[id] = (max_a_value, "word", "answer")

                    else:
                        i_values = [value for key, value in introduction.items() if key in intersection_introduction]
                        a_values = [value for key, value in answer.items() if key in intersection_answer]

                        # 각 딕셔너리에서 가장 큰 밸류의 키 확인
                        max_i_value = max(set(i_values))
                        max_a_value = max(set(a_values))

                        if max_a_value < max_i_value:
                            print(f"matching {id}, introduction")
                            id_list_select[id] = (max_i_value, "word", "introduction")                                
                        elif max_a_value > max_i_value:
                            print(f"matching {id}, answer")
                            id_list_select[id] = (max_i_value, "word", "answer")
                        else:
                            print(f"matching {id}, introduction, answer")
                            id_list_select[id] = (max_i_value, "word", "both")
        
        if id_list_select:
            trigram_exist = False
            bigram_exist = False
            for key, value in id_list_select.items():
                print(f"value: {value}")
                if value[1] == "trigram":
                    # print("got trigram")
                    trigram_exist = True
                elif value[1] == "bigram":
                    # print("got bigram")
                    bigram_exist = True
                else:
                    # print("got word")
                    pass
            
            print(f"idistselect: {id_list_select}")
            if trigram_exist == True:
                max_value = max(value[0] for value in id_list_select.values() if value[1] == "trigram")
                max_keys = [key for key, value in id_list_select.items() if value[0] == max_value and value[1] == "trigram"]


            elif bigram_exist == True:
                max_value = max(value[0] for value in id_list_select.values() if value[1] == "bigram")
                max_keys = [key for key, value in id_list_select.items() if value[0] == max_value and value[1] == "bigram"]


            else:
                max_value = max(value[0] for value in id_list_select.values() if value[1] == "word")
                max_keys = [key for key, value in id_list_select.items() if value[0] == max_value and value[1] == "word"]

            
            print(f"max_keys: {max_keys}")
            candidate_sentences = []
            if max_keys:
                for key in max_keys:
                    word_similarities = {}
                    c.execute(f"SELECT introduction, answer, i_bigrams, a_bigrams, i_trigrams, a_trigrams FROM conversations WHERE id = '{key}'")
                    results = c.fetchall()
                    for result in results:
                        introduction, answer, i_bigram, a_bigram, i_trigram, a_trigram = result
                        introduction = ast.literal_eval(introduction).keys()
                        answer = ast.literal_eval(answer).keys()
                        i_bigram = ast.literal_eval(i_bigram)
                        i_trigram = ast.literal_eval(i_trigram)
                        a_bigram = ast.literal_eval(a_bigram)
                        a_trigram = ast.literal_eval(a_trigram)

                        # 키워드에 대해 유사단어 10개 유사도와 함께 리스트에 저장
                        for keyword in word_list:
                            if keyword in model.wv:
                                similar_words = model.wv.most_similar(keyword, topn=10)
                                for word, similarity in similar_words:
                                    if word in word_similarities:
                                        word_similarities[word] += similarity  # 겹치는 단어는 유사도 합산
                                    else:
                                        word_similarities[word] = similarity

                    # 유사도가 가장 높은 10개 단어 선정
                    top_words = sorted(word_similarities, key=word_similarities.get, reverse=True)[:10]
                    print(f"generated keywords: {top_words}")

                    # a,i_bigram과 a,i_trigram 포함단어 검색
                    matching_keys_bigrams = [key for key in list(a_bigram.keys()) + list(i_bigram.keys()) if any(word in key for word in top_words)]
                    matching_keys_trigrams = [key for key in list(a_trigram.keys()) + list(i_trigram.keys()) if any(word in key for word in top_words)]
                    
                    if matching_keys_bigrams and not matching_keys_trigrams:
                        generated_sentence = reconstruct_text(matching_keys_bigrams)
                    elif matching_keys_trigrams and not matching_keys_bigrams:
                        generated_sentence = reconstruct_text(matching_keys_trigrams)
                    elif matching_keys_bigrams and matching_keys_trigrams:
                        generated_sentence = reconstruct_text(matching_keys_bigrams + matching_keys_trigrams)
                    else:
                        generated_sentence = None

                    candidate_sentences = []
                    if generated_sentence:
                        natural_sentence = ' '.join(generated_sentence)
                        # top_words에 포함된 단어만 남김
                        natural_sentence = ' '.join([word for word in natural_sentence.split() if word in top_words])
                        candidate_sentences.append(natural_sentence)


            if candidate_sentences:
                    print(f"candidate_sentences: {candidate_sentences}")
                    await ctx.channel.send(random.choice(candidate_sentences))
            else:
                # 빈문장일시 유사단어 출력
                print(top_words[0] if top_words else "No words found.")
                if top_words:
                    await ctx.channel.send(top_words[0])
    # print("\n")
    await bot.process_commands(ctx)


def reconstruct_text(ngrams):
    # 시작점 (n-gram 첫요소 시작)
    start = ngrams[0][0]
    sentence = [start]

    used_ngrams = []

    # n-gram으로 문장구성
    for i in range(len(ngrams)):
        for ngram in ngrams:
            if ngram[0] == sentence[-1] and ngram not in used_ngrams:
                sentence.append(ngram[1])
                used_ngrams.append(ngram)
                break

    return sentence


# 폰트 경로 지정
path_nanum = 'C:/Users/heake/AppData/Local/Microsoft/Windows/Fonts/NanumBarunGothic.ttf'
path_arial = 'C:/Windows/Fonts/Arial.ttf'

# 폰트 프로퍼티 설정
fontprop_nanum = fm.FontProperties(fname=path_nanum, size=18)
fontprop_arial = fm.FontProperties(fname=path_arial, size=18)
mpl.rcParams['axes.unicode_minus'] = False

def show_tsne(X, vocab):
    print(f"X:{X}, vocab:{vocab}")
    if len(X) > 1:
        tsne = TSNE(n_components=3, perplexity=max(1, len(X) - 1))
        X = tsne.fit_transform(X)
    else:
        return 0

    df = pd.DataFrame(X, index=vocab, columns=['x', 'y', 'z'])
    fig = plt.figure(figsize=(10, 5))
    ax = fig.add_subplot(111, projection='3d')
    ax.scatter(df['x'], df['y'], df['z'], color='green')

    for word, pos in df.iterrows():
        # 한글과 영어에 대해 다른 폰트 사용
        if re.search('[a-z]', word):
            ax.text(pos['x'], pos['y'], pos['z'], word, fontsize=10, fontproperties=fontprop_arial)
        else:
            ax.text(pos['x'], pos['y'], pos['z'], word, fontsize=10, fontproperties=fontprop_nanum)

    ax.set_xlabel("X", fontproperties=fontprop_arial, fontsize = 20)
    ax.set_ylabel("Y", fontproperties=fontprop_arial, fontsize = 20)
    ax.set_zlabel("Z", fontproperties=fontprop_arial, fontsize = 20)
    plt.show()



def remove_stopwords(word):
    stopwords = ['은', '는', '이', '가', '을', '를']  # 불용어 목록
    exceptions = ['있는', '요가']  # 예외 단어 목록
    if word not in exceptions:
        for stopword in stopwords:
            if word.endswith(stopword):
                return word[:-len(stopword)]
    return word

def reconstruct_sentence(bigrams, trigrams):
    # 시작점 (트라이그램의 첫 요소 시작)
    start = trigrams[0][0]
    sentence = [start]

    used_bigrams = []
    used_trigrams = []

    # 이전 트라이그램 마지막 요소 이용해서 연결
    for i in range(len(trigrams)):
        for trigram in trigrams:
            if trigram[0] == sentence[-1] and trigram not in used_trigrams:
                sentence.append(trigram[1])
                used_trigrams.append(trigram)
                break

    # 이전바이그램 마지막 요소 이용해서 연결
    for i in range(len(bigrams)):
        for bigram in bigrams:
            if bigram[0] == sentence[-1] and bigram not in used_bigrams:
                sentence.append(bigram[1])
                used_bigrams.append(bigram)
                break

    return sentence

@bot.command()
@commands.has_permissions(administrator = True)
async def train(ctx):
    await ctx.send("learning from collected data!")
    c.execute("SELECT id, word_index, introduction, i_bigrams, i_trigrams, answer, a_bigrams, a_trigrams FROM conversations")
    rows = c.fetchall()
    columns = [column[0] for column in c.description]
    dict_rows = [dict(zip(columns, row)) for row in rows]
    print(dict_rows)
    model = Word2Vec(vector_size=100, window=5, min_count=1, workers=4, sg=0)

    word_freq = Counter()
    sentences = []  # 모든 바이그램과 트라이그램을 저장할 리스트
    # 각 행의 word_index를 단어 리스트로 변환하고 모델 학습
    for row in dict_rows:
        introduction_dict = ast.literal_eval(row['introduction'])
        i_bigram_dict = ast.literal_eval(row['i_bigrams'])
        i_trigram_dict = ast.literal_eval(row['i_trigrams'])

        answer_dict = ast.literal_eval(row['answer'])
        a_bigram_dict = ast.literal_eval(row['a_bigrams'])
        a_trigram_dict = ast.literal_eval(row['a_trigrams'])

        # 불용어 제거
        introduction_dict = {remove_stopwords(word): freq for word, freq in introduction_dict.items()}
        answer_dict = {remove_stopwords(word): freq for word, freq in answer_dict.items()}
        print(f"introduction_dict: {introduction_dict}")
        print(f"answer_dict: {answer_dict}")

        word_freq.update(introduction_dict)
        word_freq.update(answer_dict)

        # 문장 재구성
        bigram_keys = list(i_bigram_dict.keys())
        trigram_keys = list(i_trigram_dict.keys())
        while len(bigram_keys) > 0 and len(trigram_keys) > 0:
            sentence = reconstruct_sentence(bigram_keys, trigram_keys)
            sentences.append(sentence)

            # 빈도수 갱신
            for bigram in bigram_keys:
                i_bigram_dict[bigram] -= 1
                if i_bigram_dict[bigram] == 0:
                    bigram_keys.remove(bigram)

            for trigram in trigram_keys:
                i_trigram_dict[trigram] -= 1
                if i_trigram_dict[trigram] == 0:
                    trigram_keys.remove(trigram)

        bigram_keys = list(a_bigram_dict.keys())
        trigram_keys = list(a_trigram_dict.keys())
        while len(bigram_keys) > 0 and len(trigram_keys) > 0:
            sentence = reconstruct_sentence(bigram_keys, trigram_keys)
            sentences.append(sentence)

            # 빈도수 갱신
            for bigram in bigram_keys:
                a_bigram_dict[bigram] -= 1
                if a_bigram_dict[bigram] == 0:
                    bigram_keys.remove(bigram)

            for trigram in trigram_keys:
                a_trigram_dict[trigram] -= 1
                if a_trigram_dict[trigram] == 0:
                    trigram_keys.remove(trigram)

        print(f"trainmodel:{sentences}")
        print(f"word_freq: {word_freq}")

    if word_freq:
        model.build_vocab_from_freq(word_freq)
    else:
        await ctx.send("no vocabulary data available")
    if sentences:
        model.train(sentences, total_examples=model.corpus_count, epochs=model.epochs) 
    else:
        await ctx.send("no sentence data available")
    model.save("chat model")
    
    if model.wv.key_to_index:
        vocab = list(model.wv.key_to_index)
        print(f"vocabraw{vocab}")
        print(f"{[model.wv.key_to_index]}")
        X = model.wv[vocab]
        show_tsne(X, vocab)


    # # 키워드 설정
    # keyword = "롤"

    # #   키워드와 가장 연관성이 높은 10개 단어
    # similar_words = model.wv.most_similar(keyword, topn=10)
    # least_similar_words = model.wv.most_similar(negative=[keyword], topn=10)

    # print("related")
    # for word, similarity in similar_words:
    #     print(f"단어: {word}, 유사도: {similarity}")
    # print("unrelated")
    # for word, similarity in least_similar_words:
    #     print(f"단어: {word}, 유사도: {similarity}")



@bot.event
async def checkbutton_callback(interaction):
    ctx = interaction.guild.name
    servername = sheetname(ctx)
    _row=checkRow(ctx)
    for row in range(1, _row+1):
       ctx = interaction.guild.name
       if openxl[servername].cell(row, 1).value == interaction.user.name and openxl[servername].cell(row=row, column=3).value == '출석안함':            
            n = openxl[servername].cell(row=row, column=2).value + 1
            openxl[servername].cell(row=row, column=2, value=n)
            openxl[servername].cell(row=row, column=3, value = '출석완료')
            openxl[servername].cell(row=row, column=4, value = "{}년 {}월 {}일".format (datetime.now().year, datetime.now().month, datetime.now().day))
            author = interaction.user
            pfp = author.avatar
            esb = discord.Embed(description="지금",  colour=discord.Colour.blue())
            esb.set_thumbnail(url=pfp)
            esb.add_field(name="{}".format(author),value="출석함!",inline=False)          
            await interaction.response.edit_message(embed = esb)
            time.sleep(1)
            esb.set_footer(text="출석하려면 버튼을 눌러주세요")
            await interaction.edit_original_response(embed=editbnr(interaction.channel))
            break
            
       elif openxl[servername].cell(row, 1).value == interaction.user.name and openxl[servername].cell(row=row, column=3).value == '출석완료':
            embedmessage2 = discord.Embed(title="오늘 이미 출석했습니다", description=interaction.user.name)
            embedmessage2.set_footer(text="잠시뒤 메세지가 자동으로 삭제됩니다")
            await interaction.response.send_message(embed=embedmessage2, ephemeral=True, delete_after=3)
            break
            
    else:
       ctx = interaction.guild.name
       openxl[servername].cell(row= row, column=1, value=interaction.user.name)
       openxl[servername].cell(row= row, column=2, value=1)
       openxl[servername].cell(row= row, column=3, value = '출석완료')
       openxl[servername].cell(row= row, column=4, value = "{}년 {}월 {}일".format (datetime.now().year, datetime.now().month, datetime.now().day))
       author = interaction.user
       pfp = author.avatar
       esb = discord.Embed(description="지금",  colour=discord.Colour.blue())
       esb.set_thumbnail(url=pfp)
       esb.add_field(name="{}".format(author),value="신규 출석함!",inline=False)
       await interaction.response.edit_message(embed=esb)
       time.sleep(1)
       esb.set_footer(text="출석하려면 버튼을 눌러주세요")
       await interaction.edit_original_response(embed=editbnr(interaction.channel))



# @bot.event
# async def on_raw_reaction_add(payload):
#     print(payload)
#     guild = discord.utils.get(bot.guilds, id= payload.guild_id)
#     targetch = discord.utils.get(guild.channels, name = f"{bot.user.name.replace(" ","-").replace("(","-").replace(")","")}의-학습-허용")
#     print(targetch.id)
#     if targetch:
#         print("1111")
#         print(payload.message_author_id)
#         print(bot.user.id)
#         if payload.channel_id == targetch.id and payload.message_author_id == bot.user.id:
#             if payload.emoji.name=='✅':
#                 c.execute("INSERT INTO allowed_ids (allowed_id) VALUES (?)", (payload.user_id,))
#                 conn.commit()
#             else:
#                 pass
                
# @bot.event
# async def on_raw_reaction_remove(payload):
#     print(payload)
#     guild = discord.utils.get(bot.guilds, id= payload.guild_id)
#     targetch = discord.utils.get(guild.channels, name = f"{bot.user.name.replace(" ","-").replace("(","-").replace(")","")}의-학습-허용")
#     print(targetch.id)
#     if targetch:
#         print("2222")
#         print(payload.message_author_id)
#         print(bot.user.id)
#         if payload.channel_id == targetch.id:
#             if payload.emoji.name=='✅':
#                 c.execute("DELETE FROM allowed_ids WHERE allowed_id = ?", (payload.user_id,))
#                 conn.commit()
#             else:
#                 pass



count = 0
@bot.event
async def on_command_error(ctx, error):
    global count
    if isinstance(error, commands.CommandNotFound):
        count += 1
        if random.randrange(1,30) == 1:
            await ctx.reply("난 이 명령어는 몰?루")
            count = 0
        elif count == 3:
            await ctx.reply("혹시 나한테 말하는거라면 난 몰?루")
    else:
        raise error



print(f"treecommand: {bot.tree.get_commands()}")
bot.run("MTM3NDU5MzMxMzE3MjM1NzI3MQ.GfF-uH.P41ZZ37JDhzMhWctzQJ8lkHcRbYT1Ywgb2DHZI")